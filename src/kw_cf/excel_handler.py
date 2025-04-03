
from pathlib import Path
from .models import WorkFlowRule,WorkFlowRules,UnclassifiedKeywords
from openpyxl import load_workbook,Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import  Dict,Optional,Callable,cast,List
from .logger_config import logger

import numpy as np
import pandas as pd
import datetime
class ExcelHandler:
    def __init__(self,error_callback:Optional[Callable]=None):
        self.error_callback:Optional[Callable] = error_callback
    def read_rules(self, file_path: Path):
        """从Excel文件中读取分词规则，并进行去重"""
        try:
            # 默认读取分词规则sheet的分词规则列
            df = pd.read_excel(file_path, sheet_name='分词规则')
            
            # 检查是否存在分词规则列
            if '分词规则' in df.columns:
                rules = df['分词规则'].dropna().astype(str).tolist()
            else:
                # 如果没有找到分词规则列，使用第一列
                rules = df.iloc[:, 0].dropna().astype(str).tolist()            
            return rules
        except Exception as e:
            raise Exception(f"读取规则文件失败: {str(e)}")
    
    def read_keywords(self, file_path: Path):
        """从Excel文件中读取关键词，并进行去重"""
        try:
            df = pd.read_excel(file_path)
            
            # 使用第一列作为关键词列
            keywords = df.iloc[:, 0].dropna().astype(str).tolist()
                        
            return keywords
        except Exception as e:
            raise Exception(f"读取关键词文件失败: {str(e)}")
    
    def save_results(self, result_df: pd.DataFrame, output_file: Path|None=None,sheet_name:str|None=None):
        """保存分类结果到Excel文件
        
        Args:
            result_df: 包含分类结果的DataFrame
            output_file: 输出文件路径，如果为None则使用默认路径
        """
        try:            
            # 如果没有指定输出文件路径，则使用默认路径
            if output_file is None:
                # 获取当前时间作为文件名的一部分
                current_time = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
                # 设置默认输出目录
                output_dir = Path('./默认输出结果')
                # 设置默认文件名
                output_file = output_dir / f'默认输出结果_{current_time}.xlsx'
            else:
                output_file = Path(output_file)
            
            # 确保输出目录存在
            try:
                # 获取目录路径
                output_dir = output_file.parent
                # 创建目录
                output_dir.mkdir(parents=True, exist_ok=True)
                logger.debug(f"已创建或确认输出目录: {output_dir.absolute()}")
            except PermissionError:
                # 权限错误时，使用用户目录作为备选
                user_dir = Path.home()
                current_time = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
                output_file = user_dir / f"关键词分类结果_{current_time}.xlsx"
                logger.warning(f"无法创建原目录，将保存到用户目录: {output_file}")
            except Exception as e:
                # 其他错误时，保存到当前目录
                current_time = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
                output_file = Path(f"关键词分类结果_{current_time}.xlsx")
                logger.error(f"创建目录时出错: {str(e)}，将保存到当前目录: {output_file}")
            
            # 确保sheet名称存在
            try:
                if sheet_name is None:
                    sheet_name = "Sheet1"
                    logger.warning(f"未指定sheet名称，将使用默认名称: {sheet_name}")
            except Exception as e:
                raise Exception(f"获取sheet名称失败: {str(e)}")
            
            # 保存到Excel
            result_df.to_excel(output_file, index=False, sheet_name=sheet_name)
            
            return output_file
        except Exception as e:
            raise Exception(f"保存结果失败: {str(e)}")


    def read_workflow_rules(self, file_path: Path) -> WorkFlowRules:
        """读取工作流规则文件
        
        Args:
            file_path: 工作流规则文件路径
            
        Returns:
            包含各个sheet规则的字典
        """
        try:
            # 检查文件名是否符合规范
            file_name = file_path.name
            if not file_name.startswith("工作流规则_"):
                raise ValueError(f"工作流规则文件名必须以'工作流规则_'开头，当前文件名: {file_name}")
            
            # 读取Excel文件的所有sheet
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            
            # 检查是否至少有Sheet1
            if 'Sheet1' not in sheet_names:
                raise ValueError("工作流规则文件必须包含Sheet1")
            rules_data = []
            # 遍历所有sheet，读取规则
            for i, sheet_name in enumerate(sheet_names):
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                # 检查sheet是否有数据
                if df.empty:
                    continue
                # 检查必需列
                required_cols = ['分类规则', '结果文件名称']
                if i > 0:  # Sheet2及以上需要额外的列
                    required_cols.append('分类sheet名称')
                if i > 1:  # Sheet3及以上需要分类标签
                    required_cols.append('分类标签')
                if i > 2:  # Sheet4及以上需要上层分类规则
                    required_cols.append('上层分类规则')
                
                missing_cols = [col for col in required_cols if col not in df.columns]
                if missing_cols:
                    raise ValueError(f"Sheet '{sheet_name}' 缺少必需列: {', '.join(missing_cols)}")
                # 读取规则数据
                for _, row in df.iterrows():
                    # 只添加非空规则
                    if not pd.notna(row['分类规则']) or not row['分类规则'].strip():
                        continue
                    rule_data ={
                        'source_sheet_name':sheet_name,
                        'rule':row['分类规则'],
                        'output_name':row['结果文件名称'], 
                        'level': i+1
                    }
                    # 添加额外的列
                    if i > 0 and '分类sheet名称' in df.columns:
                        rule_data['classified_sheet_name'] = row['分类sheet名称']
                    if i>1 and '分类标签' in df.columns:
                        rule_data['rule_tag'] = row['分类标签']
                    if i > 2 and '上层分类规则' in df.columns:
                        rule_data['parent_rule'] = row['上层分类规则']
                    rules_data.append(WorkFlowRule(**rule_data))
                    rule_data = {}
            return WorkFlowRules(rules = rules_data)
        except Exception as e:
            raise Exception(f"读取工作流规则失败: {str(e)}")
    
    def read_keyword_file(self, file_path: Path) -> UnclassifiedKeywords:
        """读取待分类文件
        
        Args:
            file_path: 待分类文件路径
            
        Returns:
            UnclassifiedKeywords
        """
        try:
            # 检查文件名是否符合规范
            file_name = file_path.name
            if not file_name.startswith("待分类_"):
                raise ValueError(f"待分类文件名必须以'待分类_'开头，当前文件名: {file_name}")
            
            # 读取Excel文件的所有sheet
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            # 读取Excel文件
            df = pd.read_excel(file_path)
            
            # 检查是否包含关键词列
            if '关键词' not in df.columns:
                raise ValueError("待分类文件必须包含'关键词'列")
            
            # 清理数据
            keywords = cast(List[str],df['关键词'].dropna().astype(str).tolist())
            return UnclassifiedKeywords(data=keywords,
                                        level=1,
                                        source_file_name=file_name,
                                        source_sheet_name=cast(str,sheet_names[0]),
                                        error_callback=self.error_callback,
                                        parent_workflow_rule_column_name=None,
                                        parent_workflow_rule_str=None)
        except Exception as e:
            raise Exception(f"读取待分类文件失败: {str(e)}")
    def read_stage_results(self, file_path: Path) -> Dict[str,pd.DataFrame]:
        """读取分类结果文件
        
        Args:
            file_path: 分类结果文件路径
            
        Returns:
            Dict[str:pd.DataFrame]classified_sheet_name:pd.DataFrame
        """
        try:
            # 读取Excel文件的所有sheet
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            result = {}
            
            for sheet_name in sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                if df.empty:
                    continue
                if '关键词' not in df.columns:
                    raise ValueError(f"Sheet '{sheet_name}' 必须包含'关键词'列")
                result[sheet_name] = df
            return result
        except Exception as e:
            raise Exception(f"读取分类结果文件失败: {str(e)}")
    
    def read_stage_classified_sheet_name(self, file_dict: Dict[str,Path]) -> Dict[str,Dict[str,list[str]|Path]]:
        """读取分类结果文件
        
        Args:
            file_path: 分类结果文件路径
            
        Returns:
           Dict[str,Dict[str,list[str]|Path]]:Dict[output_name,Dict[file_path,classified_sheet_name]]
        """
        try:
            result = {}
            for output_name,file_path in file_dict.items():
                # 读取Excel文件的所有sheet
                excel_file = pd.ExcelFile(file_path)
                sheet_names = list(excel_file.sheet_names)
                result[output_name] = {'file_path':file_path, 'classified_sheet_name':sheet_names}
            return result
        except Exception as e:
            raise Exception(f"读取分类结果文件失败: {str(e)}")
    
    def read_stage_result(self, file_path:Path,sheet_name:str) -> pd.DataFrame:
        """读取分类结果文件
        
        Args:
            file_path: 分类结果文件路径
            sheet_name: 分类结果sheet名称
        Returns:
           pd.DataFrame
        """
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        return df

    def add_sheet_data(
        self,
        df: pd.DataFrame,
        file_path: Path,
        sheet_name: str,
        create_sheet_if_missing: bool = True
    ) -> None:
        """
        向Excel文件的指定Sheet追加数据（自动处理表头和定位）
        
        Args:
            df: 要追加的DataFrame（必须包含表头）
            file_path: Excel文件路径
            sheet_name: 目标Sheet名称
            create_sheet_if_missing: 是否自动创建新Sheet
        """
        # 确保父目录存在
        file_path.parent.mkdir(parents=True, exist_ok=True)
        
        try:
            # 加载或创建Workbook
            try:
                book = load_workbook(file_path)
            except FileNotFoundError:
                book = Workbook()
                book.remove(book.active)  # type: ignore 删除默认Sheet
            
            # Sheet存在性检查
            if sheet_name not in book.sheetnames:
                if not create_sheet_if_missing:
                    raise ValueError(f"Sheet '{sheet_name}' 不存在")
                sheet = book.create_sheet(sheet_name)
                # 新Sheet写入表头
                sheet.append(df.columns.tolist())
                start_row = 2  # 表头占第1行
            else:
                sheet = book[sheet_name]
                # 检测是否为空Sheet（兼容openpyxl的max_row误差）
                start_row = sheet.max_row + 1 if any(sheet.iter_rows()) else 1
            
            # 追加数据行
            for row in dataframe_to_rows(df, index=False, header=False):
                sheet.append(row)
            
            book.save(file_path)
        
        except Exception as e:
            raise RuntimeError(f"写入失败: {str(e)}")
    


    def add_sheet_columns(
        self,
        excel_path: Path,
        sheet_name: str,
        new_df: pd.DataFrame,
        key_mapping: dict,  # 格式: {"原表列名": "新表列名"}
        tag_mapping: dict,  # 格式: {"新表标签列名": "原表输出列名"}
        new_file_path: Path|None = None,  # 新文件路径
        keep_unmatched: bool = True  # 是否保留未匹配的行
    ) -> None:
        """
        支持列名映射的数据合并工具
        
        Args:
            excel_path: 原Excel路径
            sheet_name: 目标Sheet名称
            new_df: 包含标签的新DataFrame
            key_mapping: 新旧表匹配列映射 {"原表列": "新表列"}
            tag_mapping: 标签列映射 {"新表列": "原表输出列"}
            new_file_path: 新文件路径，如果为None则覆盖原文件
            keep_unmatched: 是否保留原表未匹配的行
            wildcard: 通配符值（默认"全"）
        """
        try:
            # 1. 读取原数据
            df_old = pd.read_excel(excel_path, sheet_name=sheet_name)
            
            # 2. 列名检查
            missing_old_keys = [k for k in key_mapping if k not in df_old.columns]
            missing_new_keys = [v for v in key_mapping.values() if v not in new_df.columns]
            missing_tags = [k for k in tag_mapping if k not in new_df.columns]
            
            if missing_old_keys:
                raise ValueError(f"原表缺少匹配列: {missing_old_keys}")
            if missing_new_keys:
                raise ValueError(f"新表缺少匹配列: {missing_new_keys}") 
            if missing_tags:
                raise ValueError(f"新表缺少标签列: {missing_tags}")

            # 3. 统一列名（临时修改新DF列名以匹配原表）
            new_df_renamed = new_df.rename(
                columns={v: k for k, v in key_mapping.items()}
            )
            
            # 4. 合并数据
            how = 'left' if keep_unmatched else 'inner'
            merged = pd.merge(
                left=df_old,
                right=new_df_renamed[list(key_mapping.keys()) + list(tag_mapping.keys())],
                on=list(key_mapping.keys()),  # 使用统一后的列名匹配
                how=how
            )
            
            # 5. 重命名标签列
            merged.rename(columns=tag_mapping, inplace=True)
            
            if new_file_path:
                merged.to_excel(new_file_path, sheet_name=sheet_name, index=False)
                return None
            # 6. 写回Excel
            with pd.ExcelWriter(
                excel_path,
                engine='openpyxl',
                mode='a',
                if_sheet_exists='replace'
            ) as writer:
                merged.to_excel(writer, sheet_name=sheet_name, index=False)
                        
        except Exception as e:
            raise RuntimeError(f"处理失败: {str(e)}")
        
    def add_sheet_columns_with_wildcard(
        self,
        excel_path: Path,
        sheet_name: str,
        new_df: pd.DataFrame,
        key_mapping: dict,  # {"原表列": "新表列"}
        tag_mapping: dict,  # {"新表列": "输出列名"}
        new_file_path: Path|None = None,  # 新文件路径
        keep_unmatched: bool = True,  # 是否保留未匹配的行
        wildcard: str = "全"  # 通配符值
    ) -> None:
        """
        支持「全」值通配的智能标签合并
        
        Args:
            excel_path: 原Excel路径
            sheet_name: 目标Sheet名
            new_df: 包含标签的新DataFrame
            key_mapping: 列映射 {"原表列": "新表列"}
            tag_mapping: 标签映射 {"新表列": "输出列名"}
            wildcard: 通配符值（默认"全"）
        """
        try:
            # 1. 读取原数据
            df_old = pd.read_excel(excel_path, sheet_name=sheet_name)
            
            # 2. 列名检查
            required_old_cols = list(key_mapping.keys())
            required_new_cols = list(key_mapping.values()) + list(tag_mapping.keys())
            
            missing_old = set(required_old_cols) - set(df_old.columns)
            missing_new = set(required_new_cols) - set(new_df.columns)
            
            if missing_old:
                raise ValueError(f"原表缺少列: {missing_old}")
            if missing_new:
                raise ValueError(f"新表缺少列: {missing_new}")

            # 3. 拆分通配记录和非通配记录
            wild_mask = np.column_stack([
                new_df[col] == wildcard for col in key_mapping.values()
            ]).any(axis=1)
            
            df_wild = new_df[wild_mask].copy()  # 包含通配符的记录
            df_exact = new_df[~wild_mask]       # 精确匹配的记录

            # 4. 处理通配记录（为每个通配列创建所有可能值）
            if not df_wild.empty:
                wild_data = []
                for _, row in df_wild.iterrows():
                    # 构建通配查询条件
                    conditions = {}
                    for old_col, new_col in key_mapping.items():
                        if row[new_col] == wildcard:
                            conditions[old_col] = df_old[old_col].unique()
                        else:
                            conditions[old_col] = [row[new_col]]
                    
                    # 生成笛卡尔积组合
                    from itertools import product
                    combinations = product(*conditions.values())
                    wild_data.extend([
                        {**dict(zip(conditions.keys(), combo)), **row[list(tag_mapping.keys())]}
                        for combo in combinations
                    ])
                
                df_wild_processed = pd.DataFrame(wild_data)
                df_exact = pd.concat([df_exact, df_wild_processed])

            # 5. 统一列名后合并
            new_df_renamed = df_exact.rename(columns={v: k for k, v in key_mapping.items()})
            merged = pd.merge(
                left=df_old,
                right=new_df_renamed[list(key_mapping.keys()) + list(tag_mapping.keys())],
                on=list(key_mapping.keys()),
                how='left'
            )
            
            # 6. 重命名标签列
            merged.rename(columns=tag_mapping, inplace=True)
            
            # 7. 写回Excel
            with pd.ExcelWriter(
                excel_path,
                engine='openpyxl',
                mode='a',
                if_sheet_exists='replace'
            ) as writer:
                merged.to_excel(writer, sheet_name=sheet_name, index=False)
                
            print(f"成功更新: {excel_path}")
            print(f"通配符规则: 新表中{list(key_mapping.values())}列值为'{wildcard}'时匹配所有值")
        
        except Exception as e:
            raise RuntimeError(f"处理失败: {str(e)}")
